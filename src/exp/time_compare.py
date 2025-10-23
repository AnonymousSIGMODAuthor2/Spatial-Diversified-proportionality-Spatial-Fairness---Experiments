import sys
import os
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))

from biased_sampling import biased_sampling
from collections import defaultdict
from typing import List, Dict, Tuple
import pandas as pd
from models import Place
from config import COMBO, NUM_CELLS, GAMMAS, DATASET_NAMES
from baseline_iadu import load_dataset, iadu, load_dataset
from hybrid_sampling import hybrid, hybrid_on_grid
from grid_iadu import grid_iadu

EXPERIMENT_NAME = "times"
SHAPES = DATASET_NAMES


def save_outputs(log: List[Dict], DECIMALS_W: int = 2, DECIMALS_TIME: int = 7):
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # ---------- Build DataFrame ----------
    df = pd.DataFrame(log)

    # Column groups
    setup_cols  = ["K", "k", "g", "W", "K/(k*g)", "K'", "G", "|CL|"]
    prep_cols   = ["exact_pss_time","grid_pss_time","hybrid_exact_pss_time","hybrid_grid_pss_time"]
    select_cols = ["exact_iadu_time","grid_iadu_time", "hybrid_exact_iadu_time","hybrid_grid_iadu_time","biased_selection_time"]
    pruning_cols = ["pruning_time"]
    total_cols   = ["exact_total_time","grid_total_time","hybrid_exact_total_time","hybrid_grid_total_time"]

    all_cols = setup_cols + prep_cols + select_cols + pruning_cols + total_cols
    for col in all_cols:
        if col not in df.columns:
            df[col] = None
    df = df[all_cols]

    # Coerce numerics safely (no integer casting that can explode)
    # Setup integers (except W)
    setup_int_cols = [c for c in setup_cols if c != "W"]
    for c in setup_int_cols:
        df[c] = pd.to_numeric(df[c], errors="coerce")

    # W as float rounded to DECIMALS_W in the stored values
    if "W" in df.columns:
        df["W"] = pd.to_numeric(df["W"], errors="coerce").round(DECIMALS_W)

    # Times/pruning/totals as floats; round stored values to DECIMALS_TIME
    time_like_cols = prep_cols + select_cols + pruning_cols + total_cols
    for c in time_like_cols:
        df[c] = pd.to_numeric(df[c], errors="coerce").round(DECIMALS_TIME)

    xlsx_name = f"{EXPERIMENT_NAME}.xlsx"
    df.to_excel(xlsx_name, index=False)

    # ---------- Styling ----------
    wb = load_workbook(xlsx_name)
    ws = wb.active

    fills = {
        "setup":   PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
        "pss":     PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
        "iadu":    PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
        "pruning": PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"),
        "total":   PatternFill(start_color="E4DFEC", end_color="E4DFEC", fill_type="solid"),
    }

    block_map = {col: "setup" for col in setup_cols}
    block_map.update({col: "pss" for col in prep_cols})
    block_map.update({col: "iadu" for col in select_cols})
    block_map.update({col: "pruning" for col in pruning_cols})
    block_map.update({col: "total" for col in total_cols})

    thin, thick = Side(style="thin"), Side(style="thick")
    center_align = Alignment(horizontal="center")

    headers = [cell.value for cell in ws[1]]
    block_ends = [i for i, h in enumerate(headers) if h in prep_cols + select_cols + total_cols]

    # Exact display formats
    int_format  = "0"  # integers for setup (except W)
    w_format    = "0." + "0"*DECIMALS_W if DECIMALS_W > 0 else "0"
    time_format = "0." + "0"*DECIMALS_TIME if DECIMALS_TIME > 0 else "0"

    def style_header_row(row_cells):
        for j, cell in enumerate(row_cells):
            block = block_map.get(cell.value)
            if block:
                cell.fill = fills[block]
            cell.font = Font(bold=True)
            cell.alignment = center_align
            if j in block_ends:
                cell.border = Border(top=thin, bottom=thin, right=thick)
            elif j == 0 or j - 1 in block_ends:
                cell.border = Border(top=thin, bottom=thin, left=thick)
            else:
                cell.border = Border(top=thin, bottom=thin)

    def style_data_row(row):
        for j, cell in enumerate(row):
            header = headers[j]
            block = block_map.get(header)
            if block:
                cell.fill = fills[block]
            cell.alignment = center_align

            # Apply number formats
            if isinstance(cell.value, (int, float)) and cell.value is not None:
                if header == "W":
                    cell.number_format = w_format
                elif header in setup_int_cols:
                    cell.number_format = int_format
                elif header in time_like_cols:
                    cell.number_format = time_format

            if j in block_ends:
                cell.border = Border(bottom=thin, right=thick)
            elif j == 0 or j - 1 in block_ends:
                cell.border = Border(bottom=thin, left=thick)
            else:
                cell.border = Border(bottom=thin)

    style_header_row(ws[1])
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        style_data_row(row)

    # Wider, header-aware column widths (avoid ###)
    MIN_WIDTH = 12
    for idx, col in enumerate(ws.columns, start=1):
        header_len = len(str(ws.cell(row=1, column=idx).value or ""))
        val_len = max(len(str(c.value)) if c.value is not None else 0 for c in col)
        width = max(MIN_WIDTH, header_len + 2, val_len + 4)
        ws.column_dimensions[get_column_letter(idx)].width = width

    wb.save(xlsx_name)
    print(
        f"Styled and saved: {xlsx_name} | "
        f"Setup ints (except W); W rounded & shown with {DECIMALS_W} dp; "
        f"times/pruning/totals rounded & shown with {DECIMALS_TIME} dp."
    )

def compute_average_log(log: Dict[Tuple[int, int, int, int], List[Dict]]) -> Dict[Tuple[int, int, int, int, int], Dict]:
    """
    Average across shapes per (K, k, g, G, K') so K' is preserved
    (three rows per combo). Only numeric fields are averaged.
    """
    avg_log: Dict[Tuple[int, int, int, int, int], Dict] = {}

    for key, rows in log.items():
        if not rows:
            continue

        # Split by K' inside each (K,k,g,G)
        by_kprime: Dict[int, List[Dict]] = defaultdict(list)
        for r in rows:
            by_kprime[r["K'"]].append(r)

        for kprime, subrows in by_kprime.items():
            out: Dict = {
                "K": key[0],
                "k": key[1],
                "g": key[2],
                "G": key[3],
                "K'": kprime,
                "W": subrows[0]["W"],
                "K/(k*g)": subrows[0]["K/(k*g)"],
            }
            # Average numeric fields
            all_fields = set().union(*[r.keys() for r in subrows])
            for fname in all_fields:
                if fname in {"K", "k", "g", "G", "K'", "W", "K/(k*g)"}:
                    continue
                vals = [r[fname] for r in subrows if isinstance(r.get(fname), (int, float))]
                if vals:
                    out[fname] = sum(vals) / len(vals)
            avg_log[(key[0], key[1], key[2], key[3], kprime)] = out

    return avg_log

def run_multiple_experiments(num_runs: int = 1):
    """
    Run the whole experiment several times and average the logs again
    across runs to reduce noise.
    """
    from copy import deepcopy
    combined_log = defaultdict(list)

    for run in range(num_runs):
        print(f"=== Run {run+1}/{num_runs} ===")
        log = defaultdict(list)

        for shape in SHAPES:
            for (K, k) in COMBO:
                for g in GAMMAS:
                    W = K / (g * k)
                    for G in NUM_CELLS:
                        try:
                            S: List[Place] = load_dataset(shape, K)
                        except FileNotFoundError:
                            continue

                        K_samples = [int(K * 0.1), 
                                    #  int(K * 0.2), 
                                    #  int(K * 0.5), 
                                    #  int(K * 0.75)
                                    ]

                        # === Baselines ===
                        R_base, score_base, sum_psS_base, sum_psR_base, t_base_prep, t_base_select = iadu(S, k, W)
                        R_grid, score_grid, sum_psS_grid, sum_psR_grid, t_grid_prep, t_grid_select, _cl = grid_iadu(S, k, W, G)
                        R_biased, score_biased, sum_psS_biased, sum_psR_biased, t_biased_select = biased_sampling(S, k, W)

                        for K_sample in K_samples:
                            R_hybrid, score_hybrid, psS_sum_hybrid, psR_sum_hybrid, t_hybrid_prep, t_hybrid_select, t_pruning_exact, W_hybrid = hybrid(S, k, K_sample, W)
                            R_hybrid_grid, score_hybrid_grid, psS_sum_hg, psR_sum_hg, t_hybrid_grid_prep, t_hybrid_grid_select, t_pruning_grid = hybrid_on_grid(S, k, G, K_sample, W)

                            log[(K, k, g, G)].append({
                                "K": K, "k": k, "g": g, "W": W,
                                "K/(k*g)": f"K/(k*{g})",
                                "K'": K_sample, "G": G,
                                
                                "exact_pss_time": t_base_prep,
                                "grid_pss_time": t_grid_prep,
                                "hybrid_exact_pss_time": t_hybrid_prep,
                                "hybrid_grid_pss_time": t_hybrid_grid_prep,
                                
                                "exact_iadu_time": t_base_select,
                                "grid_iadu_time": t_grid_select,
                                "hybrid_exact_iadu_time": t_hybrid_select,
                                "hybrid_grid_iadu_time": t_hybrid_grid_select,
                                "biased_selection_time": t_biased_select,
                                
                                "pruning_time": t_pruning_exact,
                                
                                "exact_total_time": t_base_prep + t_base_select,
                                "grid_total_time": t_grid_prep + t_grid_select,
                                "hybrid_exact_total_time": t_hybrid_prep + t_hybrid_select + t_pruning_exact,
                                "hybrid_grid_total_time": t_hybrid_grid_prep + t_hybrid_grid_select + t_pruning_grid,
                            })

        # First averaging: across shapes
        avg_log = compute_average_log(log)

        # Merge into combined_log for second averaging across runs
        for key, row in avg_log.items():
            combined_log[key].append(deepcopy(row))

    # Second averaging: across runs
    final_log = {}
    for key, rows in combined_log.items():
        out = dict(rows[0])  # copy metadata
        for fname in rows[0].keys():
            if isinstance(rows[0].get(fname), (int, float)):
                vals = [r[fname] for r in rows if isinstance(r.get(fname), (int, float))]
                if vals:
                    out[fname] = sum(vals) / len(vals)
        final_log[key] = out

    save_outputs(list(final_log.values()))

# --- NEW: raw log flattener ---
def compute_raw_log(log: Dict[Tuple[int, int, int, int], List[Dict]]) -> List[Dict]:
    """
    Flatten the nested log into a plain list of rows (no averaging).
    Keeps every entry produced for each shape/K' exactly as measured.
    Output schema matches save_outputs expectations.
    """
    raw_rows: List[Dict] = []
    for _key, rows in log.items():
        for r in rows:
            raw_rows.append(dict(r))  # shallow copy
    return raw_rows


if __name__ == "__main__":
    run_multiple_experiments(num_runs=1)
