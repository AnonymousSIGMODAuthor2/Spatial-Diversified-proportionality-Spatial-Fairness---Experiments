from collections import defaultdict
from typing import List, Dict
import pandas as pd
from models import Place, SquareGrid
from config import COMBO, NUM_CELLS
from baseline_iadu import load_dataset, preparation
from grid_iadu import grid_precompute, map_place_to_cell

# ==== EXPERIMENT CONFIG ====
EXPERIMENT_NAME = "grid_pss_VS_baseline_pss"
SHAPES = ["flower", "bubble", "s_curve", "figure_eight"]


def run_experiment():
    log = defaultdict(list)

    for (K, k) in COMBO:
        W = K / k
        print(f"Running PSS comparison for K={K}, k={k}, W={W:.2f}")

        for shape in SHAPES:
            print(f"  Dataset: {shape}")
            for G in NUM_CELLS:
                print(f"  Dataset: {shape}, G={G}")
                try:
                    S: List[Place] = load_dataset(shape, K, k, G)
                except FileNotFoundError:
                    print(f"    ✗ Missing dataset: {shape}_K{K}_k{k}_G{G}.pkl")
                    continue

                # --- Baseline Preparation + IAdU ---
                psS_base, sS_base, t_base_prep = preparation(S)

                grid = SquareGrid(S, G)
                CL = list(grid.get_grid().values())
                sS_grid, psS_grid, t_grid_prep = grid_precompute(CL, S)

                place_to_cell = map_place_to_cell(CL)
                for p in S:
                    p.score = psS_grid[place_to_cell[p.id]]

                approx_error = sum(
                    100 * abs(p.score - psS_base[p.id]) / psS_base[p.id] for p in S
                ) / len(S)

                speedup_pss = t_base_prep / t_grid_prep

                log[(K, k, G)].append({
                    "|CL|": len(CL),
                    "baseline_psS_sum": sum(psS_base.values()),
                    "grid_psS_sum": sum(psS_grid.values()),
                    "approximation_error%": approx_error,
                    "baseline_pss_time": t_base_prep,
                    "grid_pss_time": t_grid_prep,
                    "speedup_psS": speedup_pss,
                })

    final_log = []
    for (K, k, G), rows in log.items():
        count = len(rows)
        accum = {}
        for row in rows:
            for key, val in row.items():
                accum[key] = accum.get(key, 0) + val

        avg_row = {key: accum[key] / count for key in accum}
        avg_row.update({
            "K": K,
            "k": k,
            "W": K / k,
            "G": G
        })
        avg_row["|CL|"] = round(avg_row["|CL|"])

        final_log.append(avg_row)

    save_outputs(final_log)


def save_outputs(log: List[Dict]):
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # --- Smart rounding ---
    def smart_round(value):
        if value == 0:
            return 0.0
        elif abs(value) >= 0.01:
            return round(value, 3)
        else:
            return float(f"{value:.5f}") if abs(value) >= 1e-5 else f"{value:.1e}"

    for row in log:
        for k, v in row.items():
            if isinstance(v, float):
                row[k] = smart_round(v)
                
    # Save raw DataFrame
    df = pd.DataFrame(log)

    # Reorder columns: priority first
    priority_cols = ["K", "k", "W", "G", "|CL|"]
    remaining_cols = [col for col in df.columns if col not in priority_cols]
    df = df[priority_cols + remaining_cols]

    xlsx_name = f"{EXPERIMENT_NAME}.xlsx"
    df.to_excel(xlsx_name, index=False)

    # Load workbook
    wb = load_workbook(xlsx_name)
    ws = wb.active

    # Setup styles
    fills = {
        "flower": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        "bubble": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        "s_curve": PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"),
        "figure_eight": PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
        "hpfr": PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid"),
        "time": PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid"),
    }

    thin = Side(style="thin")
    thick = Side(style="thick")
    regular_border = Border(top=thin, bottom=thin)
    block_border_left = Border(left=thick, top=thin, bottom=thin)
    block_border_right = Border(right=thick, top=thin, bottom=thin)
    center_align = Alignment(horizontal="center")

    # Header structure & block edges
    headers = [cell.value for cell in ws[1]]
    block_ends = []
    for i, h in enumerate(headers):
        if h in ["G", "approximation_error%", "hpfr_error%"]:
            block_ends.append(i)

    def style_header_row(row_cells):
        for j, cell in enumerate(row_cells):
            cell.font = Font(bold=True)
            cell.alignment = center_align
            if "hpfr" in str(cell.value).lower():
                cell.fill = fills["hpfr"]
            elif "time" in str(cell.value).lower():
                cell.fill = fills["time"]
            if j in block_ends:
                cell.border = block_border_right
            elif j == 0 or j - 1 in block_ends:
                cell.border = block_border_left
            else:
                cell.border = regular_border

    def style_data_row(row):
        shape = row[0].value
        fill = fills.get(shape)
        for j, cell in enumerate(row):
            cell.alignment = center_align
            if fill:
                cell.fill = fill
            if j in block_ends:
                cell.border = block_border_right
            elif j == 0 or j - 1 in block_ends:
                cell.border = block_border_left
            else:
                cell.border = regular_border

    # Style original header
    style_header_row(ws[1])

    # Style and insert headers every 25 rows (bottom-up to preserve data)
    row_count = ws.max_row
    for i in range(row_count, 2, -1):
        if (i - 2) % 25 == 0:
            ws.insert_rows(i)
            for col_idx, header_val in enumerate(headers, start=1):
                cell = ws.cell(row=i, column=col_idx, value=header_val)
            style_header_row(ws[i])


        # Auto-size columns (optional)
        for col in ws.columns:
            max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max_length + 2

    wb.save(xlsx_name)
    print(f"Saved: {EXPERIMENT_NAME}.json and .xlsx with headers every 25 lines.")


if __name__ == "__main__":
    run_experiment()
