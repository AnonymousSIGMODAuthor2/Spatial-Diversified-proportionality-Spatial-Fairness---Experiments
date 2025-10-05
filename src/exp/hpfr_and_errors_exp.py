import sys
import os
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))

from biased_sampling import biased_sampling_div
from collections import defaultdict
from typing import List, Dict, Tuple
import pandas as pd
from models import Place
from config import COMBO, NUM_CELLS, GAMMAS, DATASET_NAMES
from baseline_iadu import iadu_div, load_dataset
from hybrid_sampling import hybrid_div, hybrid_on_grid_div
from grid_iadu import base_iadu_on_grid, grid_iadu_div
import matplotlib.pyplot as plt
from matplotlib.backends.backend_pdf import PdfPages
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

EXPERIMENT_NAME = "hpfr_divided"
SHAPES = DATASET_NAMES

def run_experiment():
    log = defaultdict(list)

    for (K, k) in COMBO:
        for g in GAMMAS:
            W = K / (g * k)
            print(f"Comparing ALL methods on HPFR and times | K={K}, k={k}, g={g}, W={W:.2f}")

            for shape in SHAPES:
                for G in NUM_CELLS:
                    print(f"  Shape={shape}, G={G}")
                    
                    S: List[Place] = load_dataset(shape, K)
                    

                    # --- use three K' values ---
                    K_samples = [int(K * 0.2)]

                    # Baselines / Grid / Biased (do once per dataset)
                    R_base, score_base, score_rf_base, score_ps_base, base_pss_sum, base_psr_sum, t_base_prep, t_base_select = iadu_div(S, k, W)
                    R_grid, score_grid, score_rf_grid, score_ps_grid, grid_pss_sum, gridiadu_psr_sum, t_grid_prep, t_grid_select, lenCL = grid_iadu_div(S, k, W, G)

                    diff_grid_hpfr = pct_diff(score_grid, score_base)
                    diff_grid = pct_diff(score_grid, score_base)
                    
                    for K_sample in K_samples:
                        # Hybrids depend on K'
                        R_hybrid, score_hybrid, score_rf_hybrid, score_ps_hybrid, hybrid_pss_sum, hybrid_psr_sum, t_hybrid_prep, t_hybrid_select, t_pruning_ex, W_hybrid = hybrid_div(S, k, K_sample, W)
                        R_hybrid_grid, score_hybrid_grid, score_rf_hybrid_grid, score_ps_hybrid_grid, hybridgrid_pss_sum, hybridgrid_psr_sum, t_hybrid_grid_prep, t_hybrid_grid_select, t_pruning_grid = hybrid_on_grid_div(S, k, G, K_sample, W)

                        
                        diff_hybrid_hpfr = pct_diff(score_hybrid, score_base)
                        # Removed diff_hybrid_rf and diff_hybrid_ps
                        
                        # HYBRID GRID DIFFS
                        diff_hybrid_grid_hpfr = pct_diff(score_hybrid_grid, score_base)
                        # Removed diff_hybrid_grid_rf and diff_hybrid_grid_ps
                        
                        
                        diff_hybrid_pss_error = pct_diff(hybrid_pss_sum, base_pss_sum)
                        diff_hybrid_grid_pss_error = pct_diff(hybridgrid_pss_sum, base_pss_sum)

                        log[(K, k, g, G)].append({
                            "shape": shape,
                            "K": K,
                            "k": k,
                            "W": W,
                            "K'": K_sample,
                            "W'": W_hybrid,
                            "K/(k*g)": f"K/(k * {g})",
                            "G": G,

                            # BASELINE
                            "baseline_hpfr": score_base,
                            "baseline_score_rf": score_rf_base,
                            "baseline_score_ps": score_ps_base,
                            # Removed baseline_psS_sum
                            "baseline_psR_sum": base_psr_sum,
                            
                            # GRID IADU
                            "grid_iadu_hpfr": score_grid,
                            "grid_iadu_score_rf": score_rf_grid,
                            "grid_iadu_score_ps": score_ps_grid,
                            # Removed gridiadu_psS_sum
                            "gridiadu_psR_sum": gridiadu_psr_sum,
                            "grid_vs_base_hpfr_%": diff_grid_hpfr,
                            # Removed grid_vs_base_pss_error_%
                            
                            # HYBRID
                            "hybrid_hpfr": score_hybrid,
                            "hybrid_score_rf": score_rf_hybrid,
                            "hybrid_score_ps": score_ps_hybrid,
                            "hybrid_vs_base_hpfr_diff_%": diff_hybrid_hpfr,
                            # Removed hybrid_psS_sum
                            "hybrid_psR_sum": hybrid_psr_sum,
                            
                            # HYBRID GRID
                            "hybrid_grid_hpfr": score_hybrid_grid,
                            "hybrid_grid_score_rf": score_rf_hybrid_grid,
                            "hybrid_grid_score_ps": score_ps_hybrid_grid,
                            "hybridgrid_vs_base_hpfr_diff_%": diff_hybrid_grid_hpfr,
                            # Removed hybridgrid_psS_sum
                            "hybridgrid_psR_sum": hybridgrid_psr_sum,
                        })
                        
    avg_log = compute_average_log(log)
    save_outputs(avg_log)

import folium
from folium.plugins import MarkerCluster

def plot_folium_map(shape, K, k, G, configs: Dict[str, List[Place]], S: List[Place]):
    
    os.makedirs("maps", exist_ok=True)
    m = folium.Map(location=[0, 0], zoom_start=2, tiles="CartoDB positron")
    
    # Add full dataset S with more visible light blue markers
    for p in S:
        lat, lon = float(p.coords[0]), float(p.coords[1])
        folium.CircleMarker(
            location=[lat, lon],
            radius=3,  # larger size
            color="#3399FF",  # stronger blue
            fill=True,
            fill_color="#3399FF",
            fill_opacity=0.7,
            weight=0.8,
            popup=f"S point (id: {p.id})"
        ).add_to(m)

    colors = {
        "Baseline IAdU": "blue",
        "Grid pss + Base IAdU": "green",
        "Grid IAdU": "purple",
        "Hybrid Sampling": "orange",
        "Hybrid + Grid": "red",
        "Biased Sampling": "black"
    }

    for label, places in configs.items():
        for p in places:
            lat, lon = float(p.coords[0]), float(p.coords[1])
            folium.CircleMarker(
                location=[lat, lon],
                radius=2.5,
                color=colors.get(label, "gray"),
                fill=True,
                fill_opacity=0.6,
                popup=f"{label} (id: {p.id})"
            ).add_to(m)

    from branca.element import Template, MacroElement

    legend_html = """
    {% macro html(this, kwargs) %}
    <div style="
        position: fixed;
        bottom: 30px;
        left: 30px;
        width: 180px;
        height: 180px;
        z-index:9999;
        background-color: white;
        border:2px solid grey;
        padding: 10px;
        font-size:14px;
    ">
    <b>Legend</b><br>
    <span style="color:#3399FF;">&#9679;</span> All Places (S)<br>
    <span style="color:blue;">&#9679;</span> Baseline IAdU<br>
    <span style="color:green;">&#9679;</span> Grid pss + Base IAdU<br>
    <span style="color:purple;">&#9679;</span> Grid IAdU<br>
    <span style="color:orange;">&#9679;</span> Hybrid Sampling<br>
    <span style="color:red;">&#9679;</span> Hybrid + Grid<br>
    <span style="color:black;">&#9679;</span> Biased Sampling<br>
    </div>
    {% endmacro %}
    """
    legend = MacroElement()
    legend._template = Template(legend_html)
    m.get_root().add_child(legend)

    filename = f"maps/{shape}_K{K}_k{k}_G{G}.html"
    m.save(filename)
    print(f"✔ Saved map: {filename}")
    
def save_outputs(log: Dict):
    def smart_round(value):
        if value == 0:
            return 0.0
        elif abs(value) >= 0.01:
            return round(value, 3)
        else:
            return float(f"{value:.5f}") if abs(value) >= 1e-5 else f"{value:.1e}"

    all_rows = []
    # Note: We iterate over log.values() which contains the averaged data from compute_average_log
    for row in log.values():
        # round only numeric scalars
        for k, v in list(row.items()):
            if isinstance(v, float):
                row[k] = smart_round(v)
        all_rows.append(row)

    df = pd.DataFrame(all_rows)

    # Drop shape column — we have averages
    if "shape" in df.columns:
        df.drop(columns=["shape"], inplace=True)

    # Compute ratio only if K and K' exist
    if "K" in df.columns and "K'" in df.columns:
        df["kprime_ratio"] = df["K'"] / df["K"]
    else:
        df["kprime_ratio"] = None

    # Sort using only the columns that actually exist
    sort_cols = [c for c in ["K", "k", "K/(k*g)", "G", "kprime_ratio"] if c in df.columns]
    if sort_cols:
        df.sort_values(by=sort_cols, ascending=[True] * len(sort_cols), inplace=True)
    df.drop(columns=["kprime_ratio"], inplace=True)

    # --- Column Definitions ---
    setup_cols = ["K", "k", "W", "K'", "W'", "K/(k*g)", "G"] 
    
    # FINALIZED SCORE COLUMNS: HPFR, RF, PS scores + HPFR diff + PS R sums
    score_cols = [
        # Baseline Scores
        "baseline_hpfr", "baseline_score_rf", "baseline_score_ps", 
        
        # Grid IAdU Scores and Diff
        "grid_iadu_hpfr", "grid_iadu_score_rf", "grid_iadu_score_ps", 
        "grid_vs_base_hpfr_%", 
        
        # Hybrid Scores and Diff
        "hybrid_hpfr", "hybrid_score_rf", "hybrid_score_ps", 
        "hybrid_vs_base_hpfr_diff_%",
        
        # Hybrid Grid Scores and Diff
        "hybrid_grid_hpfr", "hybrid_grid_score_rf", "hybrid_grid_score_ps", 
        "hybridgrid_vs_base_hpfr_diff_%", 
        
        # PS R Sums (only PS R sums remaining)
        "baseline_psR_sum",
        "gridiadu_psR_sum",
        "biased_psR_sum",
        "hybrid_psR_sum",
        "hybridgrid_psR_sum",
    ]

    # Reorder columns: setup followed by scores
    all_cols = setup_cols + score_cols
    for col in all_cols:
        if col not in df.columns:
            df[col] = None
    df = df[all_cols]

    xlsx_name = f"{EXPERIMENT_NAME}.xlsx"
    df.to_excel(xlsx_name, index=False)

    # === styling (unchanged from original) ===
    wb = load_workbook(xlsx_name)
    ws = wb.active
    header_fill = PatternFill(start_color="A6A6A6", end_color="A6A6A6", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    group_fills = {
        "setup": PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"),
        "score": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
        "prep": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
        "select": PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
        "total": PatternFill(start_color="EDEDED", end_color="EDEDED", fill_type="solid"),
    }

    def apply_group_style(cols, fill, border_after=False):
        for i, col_name in enumerate(cols):
            # Check if the column exists in the final DataFrame before getting loc
            if col_name in df.columns:
                col_idx = df.columns.get_loc(col_name) + 1
                for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx, max_row=ws.max_row):
                    for cell in row:
                        cell.fill = fill
                        if border_after and i == len(cols) - 1:
                            cell.border = Border(right=Side(style='thick'))
                        else:
                            cell.border = thin_border

    apply_group_style(setup_cols, group_fills["setup"], border_after=True)
    apply_group_style(score_cols, group_fills["score"], border_after=True)

    # autosize
    from openpyxl.utils import get_column_letter
    for col in ws.columns:
        max_len = 0
        letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[letter].width = max_len + 2

    wb.save(xlsx_name)
    print(f"Results saved to {xlsx_name}")
    
def compute_average_log(
    log: Dict[Tuple[int, int, int, int], List[Dict]]
) -> Dict[Tuple[int, int, int, int, int], Dict]:
    """
    Preserve K' by averaging per (K,k,g,G,K').
    """
    from collections import defaultdict
    avg_log: Dict[Tuple[int, int, int, int, int], Dict] = {}

    for key, rows in log.items():  # key = (K, k, g, G)
        if not rows:
            continue

        by_kprime: Dict[int, List[Dict]] = defaultdict(list)
        for r in rows:
            by_kprime[r["K'"]].append(r)

        for kprime, subrows in by_kprime.items():
            # carry all setup fields explicitly (include G!)
            out = {
                "shape": subrows[0].get("shape"),
                "K": key[0],
                "k": key[1],
                "W": subrows[0]["W"],
                "K'": kprime,
                "W'": subrows[0].get("W'", None),
                "K/(k*g)": subrows[0]["K/(k*g)"],
                "G": key[3],
            }
            all_fields = set().union(*[r.keys() for r in subrows])
            for fname in all_fields:
                if fname in {"shape", "K", "k", "g", "G", "K'", "W", "W'", "K/(k*g)"}:
                    continue
                vals = [r[fname] for r in subrows if isinstance(r.get(fname), (int, float))]
                if vals:
                    out[fname] = sum(vals) / len(vals)

            avg_log[(key[0], key[1], key[2], key[3], kprime)] = out

    return avg_log

def pct_diff(new, exact):
                            return 100 * (new - exact) / exact if exact != 0 else None
                        
if __name__ == "__main__":
    run_experiment()

from collections import defaultdict

